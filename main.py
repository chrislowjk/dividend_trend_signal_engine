import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
from dateutil.relativedelta import relativedelta
import warnings
import traceback
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings('ignore')


# ------------------------------------------------------------
# Industry classification constants
# ------------------------------------------------------------
REIT_KEYWORDS = ['reit', 'trust', 'stapled']

# Business trusts and asset-heavy sectors that should use
# FFO/OCF rather than EPS for payout calculation
ASSET_HEAVY_KEYWORDS = [
    'infrastructure', 'utility', 'utilities', 'telecom',
    'transport', 'airport', 'port', 'pipeline', 'energy',
    'business trust', 'nbn tr', 'infra tr',
]


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def get_latest_value(data_series):
    """
     Return the most recent available annual value from financial data.
     """
    if data_series.empty:
        return np.nan
    values = data_series.dropna().values
    if len(values) == 0:
        return np.nan
    return values[0]


def calc_z(series, window):
    """
    Calculate rolling Z-score: (current - historical mean) / historical std dev.
    """
    mean = series.rolling(window, min_periods=int(window * 0.5)).mean()
    std  = series.rolling(window, min_periods=int(window * 0.5)).std()
    return (series - mean) / std


def find_row_keyword(df, keywords):
    """
    Search DataFrame index for keywords (case-insensitive) and return
    the first matching row.
    """
    if df.empty:
        return None

    # Create a mask on the lowercase version of the index
    mask = df.index.str.lower().str.contains('|'.join(keywords), regex=True, na=False)

    # Return the first matching row
    if mask.any():
        return df.loc[df.index[mask][0]]
    return None


def classify_industry(industry_str: str, company_name: str = '') -> dict:
    """
    Derive is_reit and is_asset_heavy flags from industry and company name.
    Returns a dict with keys: is_reit, is_asset_heavy.
    """
    ind   = str(industry_str).lower() if pd.notna(industry_str) else ''
    name  = str(company_name).lower() if pd.notna(company_name) else ''

    is_reit        = any(k in ind for k in REIT_KEYWORDS)
    is_asset_heavy = any(k in ind  for k in ASSET_HEAVY_KEYWORDS) or \
                     any(k in name for k in ASSET_HEAVY_KEYWORDS)

    # Business trusts are treated as asset-heavy
    # (e.g. NetLink NBN Trust, Keppel Infra Trust)
    if not is_reit and not is_asset_heavy:
        if 'trust' in name or 'tr' in name.split():
            is_asset_heavy = True

    return {'is_reit': is_reit, 'is_asset_heavy': is_asset_heavy}


# ------------------------------------------------------------
# 1. Fundamental Metrics (EPS / FFO / OCF)
# ------------------------------------------------------------
def get_fundamental_metrics(ticker, is_reit=False, is_asset_heavy=False):
    """
    Compute the appropriate per-share metric for payout analysis.
    Uses Funds From Operations (FFO) / Operating Cash Flow (OCF) for REITs or asset-heavy entities
    and EPS (with fallbacks) for regular companies.
    """
    try:
        financials = ticker.financials
        cashflow   = ticker.cashflow
        bs         = ticker.balance_sheet
        info       = ticker.info

        if financials.empty:
            return np.nan, 'none'

        # Get Shares Outstanding
        shares = info.get('sharesOutstanding', np.nan)

        # Fallback to Balance Sheet if info fails
        if pd.isna(shares) and not bs.empty:
            sh_row = find_row_keyword(
                bs, ['ordinary shares', 'common stock', 'share capital',
                     'units', 'equity attributable']
            )
            if sh_row is not None:
                # Balance sheet shares are a snapshot, take the most recent (first) value
                vals = sh_row.dropna().values
                if len(vals) > 0:
                    shares = vals[0]

        if pd.isna(shares) or shares == 0:
            return np.nan, 'none'

        # --- Compute FFO for REITs / Asset Heavy Companies ---
        if is_reit or is_asset_heavy:

            # Compute OCF / share
            ocf_per_share = np.nan
            if not cashflow.empty:
                ocf_row = find_row_keyword(cashflow, ['operating cash flow'])
                if ocf_row is not None:
                    ocf = get_latest_value(ocf_row)
                    if pd.notna(ocf) and ocf > 0:
                        ocf_per_share = ocf / shares

            # Get Net Income
            ni_row = find_row_keyword(financials, ['net income', 'net profit'])
            if ni_row is None:
                return np.nan, 'none'
            net_income = get_latest_value(ni_row)
            if pd.isna(net_income):
                return np.nan, 'none'

            # Compute FFO / share
            ffo_per_share = np.nan
            if not cashflow.empty:
                # Get Depreciation from Cash Flow
                dep_row = find_row_keyword(
                    cashflow, ['depreciation and amortization', 'depreciation']
                )
                dep_raw = 0.0
                if dep_row is not None:
                    v = get_latest_value(dep_row)
                    if pd.notna(v):
                        dep_raw = abs(v)

                dep_value = dep_raw if pd.notna(dep_raw) and dep_raw > 0 else 0

                # Strip disposal gains (ignore losses)
                gain_inv = 0.0
                gain_inv_row = find_row_keyword(
                    cashflow, ['gain loss on investment securities']
                )
                if gain_inv_row is not None:
                    v = get_latest_value(gain_inv_row)
                    if pd.notna(v):
                        gain_inv = max(v, 0)

                gain_bus = 0.0
                gain_bus_row = find_row_keyword(
                    cashflow, ['gain loss on sale of business']
                )
                if gain_bus_row is not None:
                    v = get_latest_value(gain_bus_row)
                    if pd.notna(v):
                        gain_bus = max(v, 0)

                # Calculate FFO
                ffo = net_income + dep_value - gain_inv - gain_bus
                if pd.notna(ffo) and ffo > 0:
                    ffo_per_share = ffo / shares

            # Use FFO when FFO and OCF are close (within 20% of each other).
            # FFO is the REIT industry standard.
            # When they diverge significantly, fall back to OCF as FFO addbacks are likely distorted
            candidates = {
                k: v for k, v in
                [('ffo', ffo_per_share), ('ocf', ocf_per_share)]
                if pd.notna(v) and v > 0
            }
            if not candidates:
                return np.nan, 'none'

            ffo_val = candidates.get('ffo', np.nan)
            ocf_val = candidates.get('ocf', np.nan)

            if pd.notna(ffo_val) and pd.notna(ocf_val):
                divergence = abs(ffo_val - ocf_val) / ocf_val
                if divergence <= 0.20:
                    # Use FFO if FFO and OCF aligns
                    return ffo_val, 'ffo'
                else:
                    # Use OCF if divergence
                    return ocf_val, 'ocf'
            else:
                # if only one metric available, use it
                best = next(iter(candidates))
                return candidates[best], best

        # --- Compute EPS for Regular Companies ---
        else:
            # Prioritize diluted EPS for conservative valuation
            eps_row = find_row_keyword(financials, ['diluted eps', 'basic eps'])
            if eps_row is not None:
                eps = get_latest_value(eps_row)
                if pd.notna(eps) and eps > 0:
                    return eps, 'eps'

            # Alternatively Net Income / Shares
            ni_row = find_row_keyword(
                financials, ['net income common stockholders', 'net income']
            )
            if ni_row is not None:
                ni = get_latest_value(ni_row)
                if pd.notna(ni) and ni > 0:
                    return ni / shares, 'eps'

            # Fall back to OCF / Shares
            if not cashflow.empty:
                ocf_row = find_row_keyword(cashflow, ['operating cash flow'])
                if ocf_row is not None:
                    ocf = get_latest_value(ocf_row)
                    if pd.notna(ocf) and ocf > 0:
                        return ocf / shares, 'ocf'

            return np.nan, 'none'

    except Exception:
        return np.nan, 'none'


# ------------------------------------------------------------
# 2. Dividend Metrics
# ------------------------------------------------------------
def compute_dividend_metrics(ticker, prices, is_reit=False,
                             is_asset_heavy=False):
    """
    Process dividend history to generate forward-looking yield estimates,
    detect dividend cuts and assess payout sustainability
    """
    # Get historical dividend data
    divs = ticker.dividends

    # Initialize DataFrame
    df = pd.DataFrame(index=prices.index)
    df['forward_div_yield']  = 0.0
    df['yield_z_blended']    = np.nan
    df['payout_ratio']       = np.nan
    df['metric_used']        = 'none'
    df['div_cut_flag']       = 0
    df['yield_trap']         = 0
    df['payout_unsustainable'] = 0

    if divs.empty:
        return df

    try:
        # Convert dividend Series to DataFrame and clean datetime index
        divs = divs.reset_index().rename(columns={'Dividends': 'div'})
        divs['Date'] = pd.to_datetime(divs['Date']).dt.tz_localize(None)
        divs = divs.set_index('Date')

        # Detect Special Dividends
        # Calculate rolling median of dividend payments over the last 4 periods to establish typical dividend amount
        divs['roll_median'] = divs['div'].rolling(4, min_periods=1).median()
        divs['roll_std']    = divs['div'].rolling(4, min_periods=1).std().fillna(0)

        # Flag special dividends if > 3x median OR > median + 2 std devs (whichever is stricter)
        threshold     = divs['roll_median'] * 3.0
        threshold_std = divs['roll_median'] + 2.0 * divs['roll_std']
        combined_threshold = threshold.combine(threshold_std, min)
        divs['special_div_flag'] = ((divs['div'] > combined_threshold) & (divs['roll_median'] > 0)).astype(int)

        # Separate regular dividends by removing special dividends
        divs['div_regular'] = divs['div'].where(divs['special_div_flag'] == 0, np.nan)

        # Reindex regular dividends to daily index, filling non-dividend days with 0
        daily_regular = divs['div_regular'].reindex(prices.index).fillna(0)

        # Calculate Trailing 12-Month (TTM) dividend
        df['annual_div_raw'] = daily_regular.rolling(252, min_periods=1).sum()

        # Smooth the TTM dividend using 252-day EWMA to reduce step-function behavior
        df['annual_div_smoothed'] = df['annual_div_raw'].ewm(span=252, adjust=False).mean()

        # Calculate forward yield as smoothed annual dividend divided by current price
        df['forward_div_yield'] = np.where(prices['Adj Close'] > 0,
                                           (df['annual_div_smoothed'] / prices['Adj Close']) * 100,
                                           0.0)

        # --- Yield Z-score ---
        # Calculate Z-scores for 5-year and 10-year windows
        lookback_5y  = 5  * 252
        lookback_10y = 10 * 252
        y_z5  = calc_z(df['forward_div_yield'], lookback_5y)
        y_z10 = calc_z(df['forward_div_yield'], lookback_10y)

        # Blend them (60% weight to 5-year, 40% to 10-year) for balanced perspective
        df['yield_z_blended'] = 0.6 * y_z5 + 0.4 * y_z10

        # To detect dividend cut, compare current TTM dividend to TTM dividend from one year ago
        ttm_prev  = df['annual_div_raw'].shift(252)
        yoy_ratio = (df['annual_div_raw'] / ttm_prev.replace(0, np.nan)).fillna(1.0)

        # Div Cut Risk: Flag dividend cut if more than 20% drop in TTM dividend
        df['div_cut_flag'] = (yoy_ratio < 0.80).astype(int)
        df['yoy_div_ratio'] = yoy_ratio

        # --- Payout ratio ---
        # Use raw TTM dividends for payout calculation
        latest_annual_div = df['annual_div_raw'].iloc[-1]

        # Get fundamental metric (FFO/OCF for REITs, EPS for regular stocks)
        fundamental_metric, metric_used = get_fundamental_metrics(
            ticker, is_reit=is_reit, is_asset_heavy=is_asset_heavy
        )
        df['metric_used'] = metric_used

        # Sustainability Risk: Payout ratio
        if pd.notna(fundamental_metric) and fundamental_metric > 0:
            current_payout  = latest_annual_div / fundamental_metric
            df['payout_ratio'] = current_payout

            if is_reit or is_asset_heavy:
                # REITs and business trusts: Flag if pays out more than 100% of FFO/OCF
                df['payout_unsustainable'] = int(current_payout > 1.0)
            else:
                # Regular companies: Flag if high payout and falling dividend
                df['payout_unsustainable'] = int(current_payout > 0.9)

        # --- Yield trap identification ---
        df['yield_trap'] = (
            (df['div_cut_flag'] == 1) |
            (df['payout_unsustainable'] == 1)
        ).astype(int)

        return df

    except Exception as e:
        print(f"Warning: Error calculating dividends for {ticker.ticker}: {e}")
        traceback.print_exc()
        return df


# ------------------------------------------------------------
# 3. Trend Metrics
# ------------------------------------------------------------
def compute_trend_metrics(prices):
    """
    Calculate trend-following metrics using log-linear regression over
    rolling 252-day windows.
    """
    # Convert to log prices for trend analysis
    log_price = np.log(prices.replace(0, np.nan))
    df = pd.DataFrame(index=prices.index)
    df['price'] = prices

    # Rolling window: 1 year of trading days for trend calculation
    window = 252
    n = len(log_price)

    # Initialize arrays for regression results
    slopes     = np.full(n, np.nan)
    intercepts = np.full(n, np.nan)
    r_squared  = np.full(n, np.nan)

    # Pre-calculate regression constants for efficiency
    x = np.arange(window)
    x_mean = x.mean()
    x_var  = ((x - x_mean) ** 2).sum()
    log_vals = log_price.values

    # Rolling regression: fit line to each 252-day window
    if n >= window:
        for i in range(window, n):
            # Log prices for this window
            y = log_vals[i - window:i]
            if np.any(np.isnan(y)):
                continue

            # Calculate ordinary least squares regression
            y_mean = y.mean()
            cov_xy = ((x - x_mean) * (y - y_mean)).sum()

            # Trend slope in log space
            slope     = cov_xy / x_var
            intercept = y_mean - slope * x_mean
            slopes[i]     = slope
            intercepts[i] = intercept

            # Calculate R-squared to measure trend strength
            y_pred  = intercept + slope * x
            ss_res  = ((y - y_pred) ** 2).sum()
            ss_tot  = ((y - y_mean) ** 2).sum()
            r_squared[i] = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0

    # Project trend line to current day (last point in window)
    trend_log = intercepts + slopes * (window - 1)

    # Convert back from log space
    df['trend_price'] = np.exp(trend_log)

    # Smooth the slope (63-day EWMA ≈ 3 months) for stable trend direction
    df['trend_slope'] = pd.Series(slopes, index=prices.index).ewm(span=63, adjust=False).mean()

    # Z-score: how far price deviates from trend (mean reversion signal)
    residual = log_price - trend_log
    resid_vol = residual.rolling(63).std()
    df['z_score_trend'] = residual / resid_vol.replace(0, np.nan)

    # Trend quality metrics (Higher = stronger trend)
    df['trend_quality'] = r_squared

    # Trend is ok if annualized slope > -5% (not in strong downtrend)
    df['trend_ok'] = (df['trend_slope'] * 252) > -0.05

    # Drawdown from 1-year high (risk metric)
    rolling_max    = prices.rolling(252).max()
    df['drawdown'] = (prices - rolling_max) / rolling_max

    # Not broken if drawdown less than 40% (still in recovery range)
    df['not_broken'] = df['drawdown'] > -0.40

    # 52-week high / low metrics (mean reversion signals)
    df['52w_high'] = prices.rolling(252, min_periods=1).max()
    df['52w_low'] = prices.rolling(252, min_periods=1).min()
    df['pct_from_52w_low'] = (prices - df['52w_low']) / df['52w_low']  * 100
    df['pct_from_52w_high'] = (prices - df['52w_high']) / df['52w_high'] * 100

    return df


# ------------------------------------------------------------
# 4. Signal Logic
# ------------------------------------------------------------
def get_combined_action(row):
    """
    Generate trading signal based on dividend yield and price trend
    """
    # Extract all relevant metrics from row
    yield_z      = row.get('yield_z_blended', np.nan)
    price_z      = row.get('z_score_trend', np.nan)
    trap         = row.get('yield_trap', 0)
    not_broken   = row.get('not_broken', False)
    pct_low      = row.get('pct_from_52w_low', 100)
    payout       = row.get('payout_ratio', np.nan)

    # Default return if no price trend data available
    if pd.isna(price_z):
        return "HOLD (No Data)", "NONE", 0

    # Initialize default values
    action = 'HOLD'
    signal_type = 'NONE'
    strength = 50

    # --- AVOID: yield trap ---
    if trap == 1:
        # Check which specific risk triggered the trap
        reasons = []
        if row.get('div_cut_flag', 0) == 1:
            reasons.append("Recent Cut")
        if row.get('payout_unsustainable', 0) == 1:
            payout_str = f"{payout:.1%}" if pd.notna(payout) else "N/A"
            reasons.append(f"High Payout ({payout_str})")

        action = f'⛔ AVOID - Dividend Trap: {", ".join(reasons)}'
        signal_type = 'AVOID'
        strength = 0

    # --- CAUTION: high payout ---
    elif pd.notna(payout) and payout > 0.95:
        action = f'⚠️ CAUTION - High Payout ({payout:.1%})'
        signal_type = 'SELL_WEAK'
        strength = 30

    # --- BUY signals (not yield trap) ---
    else:
        # STRONG BUY: deep value, safe dividend, near 52w low, trend intact
        if (pd.notna(yield_z) and yield_z > 1.5 and
                price_z < -1.0 and pct_low < 20 and not_broken):
            action = 'STRONG BUY - Deep Value Accumulation'
            signal_type = 'STRONG_BUY'
            strength = 90

        # BUY: attractive yield, near 52w low, trend intact
        elif (pd.notna(yield_z) and yield_z > 1.0 and
              pct_low < 30 and not_broken):
            action = 'BUY - Accumulate on Weakness'
            signal_type = 'BUY'
            strength = 70

        # CONTRARIAN: very cheap yield but trend broken
        elif (pd.notna(yield_z) and yield_z > 2.0 and
              not not_broken and pct_low < 40):
            action = 'CONTRARIAN BUY - High Safety, Broken Trend'
            signal_type = 'BUY_CONTRARIAN'
            strength = 65

        # OPPORTUNISTIC: growing dividend with attractive yield
        elif pd.notna(yield_z) and yield_z > 0.5:
            action = 'OPPORTUNISTIC BUY - Growing Yield'
            signal_type = 'BUY_OPPORTUNISTIC'
            strength = 60

        # TRIM: yield compressed, historically expensive
        elif pd.notna(yield_z) and yield_z < -1.5:
            action = 'TRIM - Yield Compressed, Historically Expensive'
            signal_type = 'TRIM'
            strength = 35

    return action, signal_type, strength


# ------------------------------------------------------------
# 5. Single-stock analysis
# ------------------------------------------------------------
def analyze_stock(stock_code, lookback_years=10, is_reit=False, is_asset_heavy=False):
    """
    Main function to analyze a stock and generate trading signals.
    """
    # Calculate date range for historical data
    end = datetime.now()
    start = end - relativedelta(years=lookback_years)

    try:
        # Ticker formatting for SGX stocks
        if '.' not in stock_code and '.SI' not in stock_code and '.US' not in stock_code:
            ticker_str = f'{stock_code}.SI'
        else:
            ticker_str = stock_code

        # Download price history
        ticker = yf.Ticker(ticker_str)
        prices = ticker.history(start=start, end=end, auto_adjust=False)

        if prices.empty or 'Adj Close' not in prices.columns:
            return None, None

        prices = prices[['Adj Close']].copy()
        prices.index = prices.index.tz_localize(None)
        prices = prices[prices['Adj Close'] > 0]

        if prices.empty:
            return None, None

        # Calculate dividend and price trend metrics
        div_metrics = compute_dividend_metrics(ticker, prices, is_reit=is_reit, is_asset_heavy=is_asset_heavy)
        trend_metrics = compute_trend_metrics(prices['Adj Close'])
        df = prices.join(div_metrics, how='left').join(trend_metrics, how='left')

        # Apply signal logic row-by-row to get action, type, strength
        results = df.apply(get_combined_action, axis=1, result_type='expand')
        df['action'] = results[0]
        df['signal_type'] = results[1]
        df['signal_strength'] = results[2]

        # Summary Statistics for latest day
        latest  = df.iloc[-1]
        summary = {
            'stock_code': stock_code,
            'current_price': round(latest['Adj Close'], 2),
            'div_yield': round(latest.get('forward_div_yield', 0), 2),
            'yield_z_blended': round(latest.get('yield_z_blended', 0), 2),
            'price_z_trend': round(latest.get('z_score_trend', 0), 2),
            'yoy_div_ratio': latest.get('yoy_div_ratio', 0),
            'pct_low': latest.get('pct_from_52w_low', 100),
            'payout_ratio': round(latest.get('payout_ratio', 0), 2) if pd.notna(latest.get('payout_ratio'))  else 'N/A',
            'payout_unsustainable': latest.get('payout_unsustainable', 0),
            'trend_status': 'Healthy' if latest.get('trend_ok') else 'Weak',
            'yield_trap': latest.get('yield_trap', 0),
            'div_cut_flag': latest.get('div_cut_flag', 0),
            'not_broken': latest.get('not_broken', False),
            'action': latest['action'],
            'signal_type': latest['signal_type'],
            'signal_strength': latest['signal_strength'],
        }
        return df, summary

    except Exception as e:
        print(f"Error processing {stock_code}: {e}")
        traceback.print_exc()
        return None, None


# ------------------------------------------------------------
# 6. Batch analysis
# ------------------------------------------------------------
def batch_analysis(stock_list, lookback_years=10):
    """
    Analyze multiple stocks and return a sorted summary DataFrame.
    """
    summaries = []
    print(f"Analyzing {len(stock_list)} stocks...")

    for num, item in enumerate(stock_list):
        stock_code   = item['code']
        company_name = item.get('name', '')
        raw_industry = item.get('industry', '')

        # Auto-classify from industry + name, then allow manual overrides
        flags = classify_industry(raw_industry, company_name)
        is_reit        = item.get('is_reit',        flags['is_reit'])
        is_asset_heavy = item.get('is_asset_heavy', flags['is_asset_heavy'])

        # Run stock analysis pipeline
        print(f"[{num + 1}/{len(stock_list)}] Processing {company_name}...", end='\r')
        df, summary = analyze_stock(stock_code, lookback_years=lookback_years, is_reit=is_reit, is_asset_heavy=is_asset_heavy)

        if summary:
            ordered_summary = {
                'stock_code':   summary['stock_code'],
                'company_name': company_name,
                'industry':     raw_industry,
                'is_reit':      is_reit,
                'is_asset_heavy': is_asset_heavy,
                **{k: v for k, v in summary.items()
                   if k not in ('stock_code', 'is_reit', 'is_asset_heavy')},
            }
            summaries.append(ordered_summary)

    print("\nAnalysis complete.")
    if not summaries:
        return pd.DataFrame()

    summary_df = pd.DataFrame(summaries)
    summary_df = summary_df.sort_values(
        by=['signal_strength', 'price_z_trend', 'yield_z_blended'],
        ascending=[False, True, True]
    )
    return summary_df


# ------------------------------------------------------------
# 8. Main execution block
# ------------------------------------------------------------
if __name__ == "__main__":

    file = 'sti_component.xlsx'
    sti_comp = pd.read_excel(file)

    stock_list = []
    for name, code, ind in zip(sti_comp['name'], sti_comp['code'], sti_comp['industry']):
        stock_list.append({'code': code, 'name': name, 'industry': ind})

    print(f"Loaded {len(stock_list)} stocks from Excel.")

    result_df = batch_analysis(stock_list, lookback_years=10)

    if not result_df.empty:
        output_filename = 'dividend_trend_signal_output.xlsx'

        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:

            # Write dataframe to sheet
            result_df.to_excel(writer, index=False, sheet_name='analysis')

            # Access the workbook and worksheet
            ws = writer.sheets['analysis']

            # Green Fill for BUY actions
            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            font_dark_green = Font(color="006100", bold=True)

            # Red Fill for AVOID/CAUTION actions
            fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            font_dark_red = Font(color="9C0006", bold=True)

            # Yellow for HOLD actions
            fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            font_dark_yellow = Font(color="9C5700", bold=True)

            # Orange for TRIM actions
            fill_orange = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
            font_dark_orange = Font(color="7F3F00", bold=True)

            # Get column index for 'action'
            try:
                action_col_idx = result_df.columns.get_loc('action') + 1
            except KeyError:
                action_col_idx = None

            # Get the action value
            for row_idx, row in enumerate(result_df.itertuples(index=False), start=2):
                action_value = row[result_df.columns.get_loc('action')] if action_col_idx else ""
                cell = ws.cell(row=row_idx, column=action_col_idx)

                # Apply Formatting based on Action Text
                if isinstance(action_value, str):
                    action_upper = action_value.upper()

                    # Buy Logic (Strong Buy, Buy, Contrarian Buy)
                    if 'BUY' in action_upper and 'AVOID' not in action_upper:
                        for col in range(1, len(result_df.columns) + 1):
                            c = ws.cell(row=row_idx, column=col)
                            c.fill = fill_green
                            c.font = font_dark_green

                    # Avoid/Caution Logic
                    elif 'AVOID' in action_upper or 'CAUTION' in action_upper:
                        for col in range(1, len(result_df.columns) + 1):
                            c = ws.cell(row=row_idx, column=col)
                            c.fill = fill_red
                            c.font = font_dark_red

                    # Hold Logic
                    elif 'HOLD' in action_upper:
                        for col in range(1, len(result_df.columns) + 1):
                            c = ws.cell(row=row_idx, column=col)
                            c.fill = fill_yellow
                            c.font = font_dark_yellow

                    # Trim Logic
                    elif 'TRIM' in action_upper:
                        for col in range(1, len(result_df.columns) + 1):
                            c = ws.cell(row=row_idx, column=col)
                            c.fill = fill_orange
                            c.font = font_dark_orange

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        print(f"Successfully saved formatted report to '{output_filename}'")
    else:
        print("No results generated.")